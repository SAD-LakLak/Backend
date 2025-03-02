from django.shortcuts import render, get_object_or_404
from django.utils import timezone
from django.db import transaction
from django.db.models import F, Count, Sum, Avg, Q
from rest_framework import viewsets, status, filters
from rest_framework.decorators import api_view, permission_classes, action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.views import APIView
from django.utils.dateparse import parse_date
from datetime import timedelta
import csv
import io
import openpyxl
from django.http import HttpResponse
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
from django.template.loader import render_to_string
# from weasyprint import HTML  # Temporarily commented out due to missing GTK dependencies
import tempfile
from django.conf import settings
import os
from django.db.models import Sum, Count, Case, When, IntegerField

from core.models import Product
from core.permissions import IsSupplier, IsPackageCombinator, IsSupervisor
from .models import InventoryTransaction, LowStockAlert, PriceChangeLog
from .serializers import (
    ProductInventorySerializer, InventoryTransactionSerializer, 
    LowStockAlertSerializer, PriceChangeLogSerializer,
    StockUpdateSerializer, BulkStockUpdateSerializer,
    PriceUpdateSerializer, BulkPriceUpdateSerializer
)
from .kafka_utils import (
    send_inventory_update, send_low_stock_alert, 
    send_price_change_event, send_product_created_event,
    send_product_deleted_event
)

class ProductInventoryViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = ProductInventorySerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['name', 'type']
    ordering_fields = ['name', 'stock', 'price', 'type']
    filterset_fields = ['is_active', 'type', 'provider']
    
    def get_queryset(self):
        user = self.request.user
        if user.role == 'supplier':
            return Product.objects.filter(provider=user)
        elif user.role in ['package_combinator', 'supervisor']:
            return Product.objects.all()
        return Product.objects.filter(is_active=True)

class InventoryTransactionViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = InventoryTransactionSerializer
    permission_classes = [IsAuthenticated & (IsSupplier | IsSupervisor)]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['product', 'transaction_type', 'performed_by']
    ordering_fields = ['timestamp', 'quantity']
    
    def get_queryset(self):
        user = self.request.user
        if user.role == 'supplier':
            return InventoryTransaction.objects.filter(product__provider=user)
        return InventoryTransaction.objects.all()

class LowStockAlertViewSet(viewsets.ModelViewSet):
    serializer_class = LowStockAlertSerializer
    permission_classes = [IsAuthenticated & (IsSupplier | IsSupervisor)]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['product', 'status']
    ordering_fields = ['created_at', 'stock_level']
    
    def get_queryset(self):
        user = self.request.user
        if user.role == 'supplier':
            return LowStockAlert.objects.filter(product__provider=user)
        return LowStockAlert.objects.all()
    
    @action(detail=True, methods=['post'])
    def acknowledge(self, request, pk=None):
        alert = self.get_object()
        if alert.status != 'pending':
            return Response(
                {"detail": "Alert is already acknowledged or resolved."},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        alert.status = 'acknowledged'
        alert.acknowledged_by = request.user
        alert.acknowledged_at = timezone.now()
        alert.save()
        
        return Response(
            {"detail": "Alert acknowledged successfully."},
            status=status.HTTP_200_OK
        )
    
    @action(detail=True, methods=['post'])
    def resolve(self, request, pk=None):
        alert = self.get_object()
        if alert.status == 'resolved':
            return Response(
                {"detail": "Alert is already resolved."},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        alert.status = 'resolved'
        alert.resolved_at = timezone.now()
        alert.save()
        
        return Response(
            {"detail": "Alert resolved successfully."},
            status=status.HTTP_200_OK
        )

class PriceChangeLogViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = PriceChangeLogSerializer
    permission_classes = [IsAuthenticated & (IsSupplier | IsSupervisor)]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['product', 'changed_by']
    ordering_fields = ['changed_at']
    
    def get_queryset(self):
        user = self.request.user
        if user.role == 'supplier':
            return PriceChangeLog.objects.filter(product__provider=user)
        return PriceChangeLog.objects.all()

@api_view(['POST'])
@permission_classes([IsAuthenticated & IsSupplier])
def update_stock(request):
    serializer = StockUpdateSerializer(data=request.data)
    if not serializer.is_valid():
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    product_id = serializer.validated_data['product_id']
    quantity = serializer.validated_data['quantity']
    notes = serializer.validated_data.get('notes', '')
    transaction_type = serializer.validated_data['transaction_type']
    
    try:
        with transaction.atomic():
            product = get_object_or_404(Product, id=product_id, provider=request.user)
            previous_stock = product.stock
            
            if transaction_type == 'add':
                product.stock = F('stock') + quantity
            elif transaction_type == 'remove':
                if product.stock < quantity:
                    return Response(
                        {"detail": "Not enough stock available."},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                product.stock = F('stock') - quantity
            elif transaction_type == 'adjust':
                product.stock = quantity
            
            product.save()
            product.refresh_from_db()
            
            InventoryTransaction.objects.create(
                product=product,
                quantity=quantity if transaction_type != 'adjust' else (quantity - previous_stock),
                previous_stock=previous_stock,
                new_stock=product.stock,
                transaction_type=transaction_type,
                notes=notes,
                performed_by=request.user
            )
            
            send_inventory_update(
                product_id=product.id,
                old_stock=previous_stock,
                new_stock=product.stock,
                user_id=request.user.id
            )
            
            return Response({
                "detail": "Stock updated successfully.",
                "product_id": product.id,
                "previous_stock": previous_stock,
                "new_stock": product.stock
            }, status=status.HTTP_200_OK)
    
    except Exception as e:
        return Response(
            {"detail": f"Error updating stock: {str(e)}"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )

@api_view(['POST'])
@permission_classes([IsAuthenticated & IsSupplier])
def bulk_update_stock(request):
    serializer = BulkStockUpdateSerializer(data=request.data)
    if not serializer.is_valid():
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    product_ids = serializer.validated_data['product_ids']
    quantity = serializer.validated_data['quantity']
    notes = serializer.validated_data.get('notes', '')
    transaction_type = serializer.validated_data['transaction_type']
    
    results = {
        'successful': [],
        'failed': []
    }
    
    for product_id in product_ids:
        try:
            with transaction.atomic():
                product = get_object_or_404(Product, id=product_id, provider=request.user)
                previous_stock = product.stock
                
                if transaction_type == 'add':
                    product.stock = F('stock') + quantity
                elif transaction_type == 'remove':
                    if product.stock < quantity:
                        results['failed'].append({
                            'product_id': product_id,
                            'reason': 'Not enough stock available'
                        })
                        continue
                    product.stock = F('stock') - quantity
                elif transaction_type == 'adjust':
                    product.stock = quantity
                
                product.save()
                product.refresh_from_db()
                
                InventoryTransaction.objects.create(
                    product=product,
                    quantity=quantity if transaction_type != 'adjust' else (quantity - previous_stock),
                    previous_stock=previous_stock,
                    new_stock=product.stock,
                    transaction_type=transaction_type,
                    notes=notes,
                    performed_by=request.user
                )
                
                send_inventory_update(
                    product_id=product.id,
                    old_stock=previous_stock,
                    new_stock=product.stock,
                    user_id=request.user.id
                )
                
                results['successful'].append({
                    'product_id': product.id,
                    'previous_stock': previous_stock,
                    'new_stock': product.stock
                })
                
        except Exception as e:
            results['failed'].append({
                'product_id': product_id,
                'reason': str(e)
            })
    
    return Response({
        "detail": "Bulk stock update completed.",
        "results": results
    }, status=status.HTTP_200_OK)

@api_view(['POST'])
@permission_classes([IsAuthenticated & IsSupplier])
def update_price(request):
    serializer = PriceUpdateSerializer(data=request.data)
    if not serializer.is_valid():
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    product_id = serializer.validated_data['product_id']
    new_price = serializer.validated_data['new_price']
    notes = serializer.validated_data.get('notes', '')
    
    try:
        with transaction.atomic():
            product = get_object_or_404(Product, id=product_id, provider=request.user)
            old_price = product.price
            
            if old_price == new_price:
                return Response({
                    "detail": "No price change needed.",
                    "product_id": product.id,
                    "price": product.price
                }, status=status.HTTP_200_OK)
            
            product.price = new_price
            product.save()
            
            PriceChangeLog.objects.create(
                product=product,
                old_price=old_price,
                new_price=new_price,
                changed_by=request.user,
                notes=notes
            )
            
            send_price_change_event(
                product_id=product.id,
                old_price=old_price,
                new_price=new_price,
                user_id=request.user.id
            )
            
            return Response({
                "detail": "Price updated successfully.",
                "product_id": product.id,
                "old_price": old_price,
                "new_price": new_price
            }, status=status.HTTP_200_OK)
    
    except Exception as e:
        return Response(
            {"detail": f"Error updating price: {str(e)}"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )

@api_view(['POST'])
@permission_classes([IsAuthenticated & IsSupplier])
def bulk_update_price(request):
    serializer = BulkPriceUpdateSerializer(data=request.data)
    if not serializer.is_valid():
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    product_ids = serializer.validated_data['product_ids']
    new_price = serializer.validated_data['new_price']
    notes = serializer.validated_data.get('notes', '')
    
    results = {
        'successful': [],
        'failed': []
    }
    
    for product_id in product_ids:
        try:
            with transaction.atomic():
                product = get_object_or_404(Product, id=product_id, provider=request.user)
                old_price = product.price
                
                if old_price == new_price:
                    results['successful'].append({
                        'product_id': product.id,
                        'message': 'No price change needed',
                        'price': product.price
                    })
                    continue
                
                product.price = new_price
                product.save()
                
                PriceChangeLog.objects.create(
                    product=product,
                    old_price=old_price,
                    new_price=new_price,
                    changed_by=request.user,
                    notes=notes
                )
                
                send_price_change_event(
                    product_id=product.id,
                    old_price=old_price,
                    new_price=new_price,
                    user_id=request.user.id
                )
                
                results['successful'].append({
                    'product_id': product.id,
                    'old_price': old_price,
                    'new_price': new_price
                })
                
        except Exception as e:
            results['failed'].append({
                'product_id': product_id,
                'reason': str(e)
            })
    
    return Response({
        "detail": "Bulk price update completed.",
        "results": results
    }, status=status.HTTP_200_OK)

class InventoryDashboardView(APIView):
    permission_classes = [IsAuthenticated & (IsSupplier | IsSupervisor)]
    
    def get(self, request):
        user = request.user
        date_from = request.query_params.get('date_from')
        date_to = request.query_params.get('date_to')
        
        transactions_filter = Q()
        if date_from:
            try:
                date_from = parse_date(date_from)
                transactions_filter &= Q(timestamp__gte=date_from)
            except:
                pass
        
        if date_to:
            try:
                date_to = parse_date(date_to)
                date_to = date_to + timedelta(days=1)
                transactions_filter &= Q(timestamp__lt=date_to)
            except:
                pass
        
        if user.role == 'supplier':
            products = Product.objects.filter(provider=user)
            transactions = InventoryTransaction.objects.filter(
                product__provider=user
            ).filter(transactions_filter)
            alerts = LowStockAlert.objects.filter(product__provider=user)
            price_changes = PriceChangeLog.objects.filter(product__provider=user)
        else:
            products = Product.objects.all()
            transactions = InventoryTransaction.objects.filter(transactions_filter)
            alerts = LowStockAlert.objects.all()
            price_changes = PriceChangeLog.objects.all()
        
        total_products = products.count()
        active_products = products.filter(is_active=True).count()
        low_stock_products = products.filter(stock__lte=settings.LOW_STOCK_THRESHOLD).count()
        out_of_stock_products = products.filter(stock=0).count()
        
        total_transactions = transactions.count()
        add_transactions = transactions.filter(transaction_type='add').aggregate(
            count=Count('id'),
            total_quantity=Sum('quantity')
        )
        remove_transactions = transactions.filter(transaction_type='remove').aggregate(
            count=Count('id'),
            total_quantity=Sum('quantity')
        )
        
        pending_alerts = alerts.filter(status='pending').count()
        acknowledged_alerts = alerts.filter(status='acknowledged').count()
        resolved_alerts = alerts.filter(status='resolved').count()
        
        total_price_changes = price_changes.count()
        avg_price_increase = price_changes.filter(
            new_price__gt=F('old_price')
        ).aggregate(
            avg_increase=Avg(F('new_price') - F('old_price'))
        )['avg_increase'] or 0
        
        avg_price_decrease = price_changes.filter(
            new_price__lt=F('old_price')
        ).aggregate(
            avg_decrease=Avg(F('old_price') - F('new_price'))
        )['avg_decrease'] or 0
        
        recent_transactions = transactions.order_by('-timestamp')[:10]
        recent_transactions_data = InventoryTransactionSerializer(recent_transactions, many=True).data
        
        recent_alerts = alerts.order_by('-created_at')[:10]
        recent_alerts_data = LowStockAlertSerializer(recent_alerts, many=True).data
        
        recent_price_changes = price_changes.order_by('-changed_at')[:10]
        recent_price_changes_data = PriceChangeLogSerializer(recent_price_changes, many=True).data
        
        return Response({
            'product_stats': {
                'total_products': total_products,
                'active_products': active_products,
                'low_stock_products': low_stock_products,
                'out_of_stock_products': out_of_stock_products,
            },
            'transaction_stats': {
                'total_transactions': total_transactions,
                'add_transactions': add_transactions,
                'remove_transactions': remove_transactions,
            },
            'alert_stats': {
                'pending_alerts': pending_alerts,
                'acknowledged_alerts': acknowledged_alerts,
                'resolved_alerts': resolved_alerts,
            },
            'price_stats': {
                'total_price_changes': total_price_changes,
                'avg_price_increase': avg_price_increase,
                'avg_price_decrease': avg_price_decrease,
            },
            'recent_data': {
                'transactions': recent_transactions_data,
                'alerts': recent_alerts_data,
                'price_changes': recent_price_changes_data,
            }
        }, status=status.HTTP_200_OK)

@api_view(['GET'])
@permission_classes([IsAuthenticated & (IsSupplier | IsSupervisor)])
def export_inventory(request):
    export_format = request.query_params.get('format', 'csv')
    user = request.user
    
    if user.role == 'supplier':
        products = Product.objects.filter(provider=user, is_deleted=False)
    else:
        products = Product.objects.filter(is_deleted=False)
    
    if export_format == 'excel':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Inventory"
        
        headers = ['ID', 'Name', 'Type', 'Price', 'Stock', 'Active', 'Provider']
        
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        for row_num, product in enumerate(products, 2):
            ws.cell(row=row_num, column=1).value = product.id
            ws.cell(row=row_num, column=2).value = product.name
            ws.cell(row=row_num, column=3).value = product.type
            ws.cell(row=row_num, column=4).value = product.price
            ws.cell(row=row_num, column=5).value = product.stock
            ws.cell(row=row_num, column=6).value = 'Yes' if product.is_active else 'No'
            ws.cell(row=row_num, column=7).value = product.provider.username
            
            if product.stock <= settings.LOW_STOCK_THRESHOLD:
                for col_num in range(1, 8):
                    cell = ws.cell(row=row_num, column=col_num)
                    if product.stock == 0:
                        cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                    else:
                        cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        
        for col_num in range(1, 8):
            ws.column_dimensions[get_column_letter(col_num)].width = 15
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=inventory.xlsx'
        
        wb.save(response)
        return response
    
    else:
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename=inventory.csv'
        
        writer = csv.writer(response)
        writer.writerow(['ID', 'Name', 'Type', 'Price', 'Stock', 'Active', 'Provider'])
        
        for product in products:
            writer.writerow([
                product.id,
                product.name,
                product.type,
                product.price,
                product.stock,
                'Yes' if product.is_active else 'No',
                product.provider.username
            ])
        
        return response

@api_view(['POST'])
@permission_classes([IsAuthenticated & IsSupplier])
def import_inventory(request):
    if 'file' not in request.FILES:
        return Response(
            {"detail": "No file provided."},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    file = request.FILES['file']
    file_extension = file.name.split('.')[-1].lower()
    
    results = {
        'successful': [],
        'failed': []
    }
    
    try:
        if file_extension == 'xlsx':
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            
            rows = list(ws.rows)[1:]
            
            for row in rows:
                try:
                    product_id = row[0].value
                    if not product_id:
                        continue
                    
                    product = get_object_or_404(
                        Product, 
                        id=product_id, 
                        provider=request.user,
                        is_deleted=False
                    )
                    
                    new_stock = int(row[4].value)
                    old_stock = product.stock
                    
                    product.stock = new_stock
                    product.save()
                    
                    transaction_type = 'adjust'
                    InventoryTransaction.objects.create(
                        product=product,
                        quantity=new_stock - old_stock,
                        previous_stock=old_stock,
                        new_stock=new_stock,
                        transaction_type=transaction_type,
                        notes='Imported from Excel',
                        performed_by=request.user
                    )
                    
                    send_inventory_update(
                        product_id=product.id,
                        old_stock=old_stock,
                        new_stock=new_stock,
                        user_id=request.user.id
                    )
                    
                    results['successful'].append({
                        'product_id': product.id,
                        'name': product.name,
                        'previous_stock': old_stock,
                        'new_stock': new_stock
                    })
                    
                except Exception as e:
                    results['failed'].append({
                        'product_id': product_id if 'product_id' in locals() else 'Unknown',
                        'reason': str(e)
                    })
        
        elif file_extension == 'csv':
            decoded_file = file.read().decode('utf-8')
            io_string = io.StringIO(decoded_file)
            reader = csv.reader(io_string)
            
            next(reader)
            
            for row in reader:
                try:
                    if not row or not row[0]:
                        continue
                    
                    product_id = int(row[0])
                    product = get_object_or_404(
                        Product, 
                        id=product_id, 
                        provider=request.user,
                        is_deleted=False
                    )
                    
                    new_stock = int(row[4])
                    old_stock = product.stock
                    
                    product.stock = new_stock
                    product.save()
                    
                    transaction_type = 'adjust'
                    InventoryTransaction.objects.create(
                        product=product,
                        quantity=new_stock - old_stock,
                        previous_stock=old_stock,
                        new_stock=new_stock,
                        transaction_type=transaction_type,
                        notes='Imported from CSV',
                        performed_by=request.user
                    )

                    send_inventory_update(
                        product_id=product.id,
                        old_stock=old_stock,
                        new_stock=new_stock,
                        user_id=request.user.id
                    )
                    
                    results['successful'].append({
                        'product_id': product.id,
                        'name': product.name,
                        'previous_stock': old_stock,
                        'new_stock': new_stock
                    })
                    
                except Exception as e:
                    results['failed'].append({
                        'product_id': row[0] if row and len(row) > 0 else 'Unknown',
                        'reason': str(e)
                    })
        
        else:
            return Response(
                {"detail": "Unsupported file format. Please upload CSV or Excel file."},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        return Response({
            "detail": "Import completed.",
            "results": results
        }, status=status.HTTP_200_OK)
        
    except Exception as e:
        return Response(
            {"detail": f"Error processing file: {str(e)}"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )

@api_view(['GET'])
@permission_classes([IsAuthenticated & (IsSupplier | IsSupervisor)])
def generate_inventory_report(request):
    user = request.user
    report_type = request.query_params.get('type', 'inventory')
    date_from = request.query_params.get('date_from')
    date_to = request.query_params.get('date_to')
    
    date_filter = Q()
    if date_from:
        try:
            date_from = parse_date(date_from)
            date_filter &= Q(timestamp__gte=date_from)
        except:
            pass
    
    if date_to:
        try:
            date_to = parse_date(date_to)
            date_to = date_to + timedelta(days=1)
            date_filter &= Q(timestamp__lt=date_to)
        except:
            pass
    
    if user.role == 'supplier':
        products = Product.objects.filter(provider=user, is_deleted=False)
        transactions = InventoryTransaction.objects.filter(
            product__provider=user
        ).filter(date_filter)
    else:
        products = Product.objects.filter(is_deleted=False)
        transactions = InventoryTransaction.objects.filter(date_filter)
    
    context = {
        'user': user,
        'generated_at': timezone.now(),
        'date_from': date_from,
        'date_to': date_to if date_to else timezone.now(),
    }
    
    if report_type == 'transactions':
        context['report_title'] = 'Inventory Transaction Report'
        context['transactions'] = transactions.order_by('-timestamp')
        
        transaction_summary = transactions.values('transaction_type').annotate(
            count=Count('id'),
            total_quantity=Sum('quantity')
        )
        context['transaction_summary'] = transaction_summary
        
        product_transactions = transactions.values('product__name').annotate(
            count=Count('id'),
            total_quantity=Sum('quantity'),
            additions=Sum(Case(
                When(transaction_type='add', then='quantity'),
                default=0,
                output_field=IntegerField()
            )),
            removals=Sum(Case(
                When(transaction_type='remove', then='quantity'),
                default=0,
                output_field=IntegerField()
            ))
        ).order_by('-count')
        context['product_transactions'] = product_transactions
        
        template_name = 'inventory/transaction_report.html'
    
    elif report_type == 'low_stock':
        low_stock_products = products.filter(stock__lte=settings.LOW_STOCK_THRESHOLD)
        context['report_title'] = 'Low Stock Report'
        context['low_stock_products'] = low_stock_products
        context['threshold'] = settings.LOW_STOCK_THRESHOLD
        
        stock_levels = low_stock_products.values('stock').annotate(
            count=Count('id')
        ).order_by('stock')
        context['stock_levels'] = stock_levels
        
        out_of_stock = low_stock_products.filter(stock=0).count()
        context['out_of_stock'] = out_of_stock
        
        template_name = 'inventory/low_stock_report.html'
    
    else:
        context['report_title'] = 'Inventory Status Report'
        context['products'] = products
        
        total_products = products.count()
        active_products = products.filter(is_active=True).count()
        low_stock_products = products.filter(stock__lte=settings.LOW_STOCK_THRESHOLD).count()
        out_of_stock_products = products.filter(stock=0).count()
        
        context['inventory_summary'] = {
            'total_products': total_products,
            'active_products': active_products,
            'low_stock_products': low_stock_products,
            'out_of_stock_products': out_of_stock_products,
        }
        
        product_types = products.values('type').annotate(
            count=Count('id'),
            total_stock=Sum('stock')
        ).order_by('type')
        context['product_types'] = product_types
        
        template_name = 'inventory/inventory_report.html'
    
    with tempfile.TemporaryDirectory() as temp_dir:
        html_string = render_to_string(template_name, context)
        response = HttpResponse(html_string, content_type='text/html')
        return response
